import functools
import multiprocessing
import openpyxl
import os
import pandas
import time
import win32com.client as win

from datetime import datetime
from multiprocessing import Pool
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from os.path import join as pjoin
from os import mkdir
from resources import src
from resources.utils import read_json_file


json_data = None


def timer(func):
    @functools.wraps(func)
    def wrapper_timer(*args, **kwargs):
        start_time = time.perf_counter()
        obj = func(*args, **kwargs)
        end_time = time.perf_counter()
        run_time = end_time - start_time
        m, s = divmod(run_time, 60)

        msg = f"\nexecute time: {int(m)} min, {int(s)} sec"
        print(msg, flush=True)

    return wrapper_timer


class ExcelHandler:
    roc = datetime.today().year - 1911
    titles = {
        'sample_info': ['農戶編號', '調查姓名', '出生年', '層別', '連結編號', '電話', '地址', ],
        'household': ['出生年', '關係', '死亡或\n除戶', '農保/農職', '老農津貼', '國保給付', '勞保給付',
                      '勞退給付', '農保給付', '住院\n日數\n(1-8月)', '門診\n次數\n(1-8月)',
                      '健保\n身分別', '健保被\n保險人\n註記', '應繳眷\n口數', '健保自\n付金額\n(1-8月)',
                      '勞保費\n8月', '國保實\n收保費\n(1-8月)', ],
        'fallow_declare': ['申報核定'],
        'fallow_transfer_subsidy': ['轉作補貼', '項目', '作物名稱', '金額', '期別'],
        'disaster_subsidy': ['災害', '項目', '災害', '核定作物', '核定面積', '金額'],
        'livestock': ['畜牧資訊', '年份', '調查時間', '畜牧品項', '在養頭數', '供應\n屠宰數', ],
        'small_large_data': ['小大補貼', '姓名', '大專業農\n轉契作', '小地主\n出租給付', '離農獎勵',
                             '期別', '是否為\n小大'],
        'crop_name': ['作物名稱'],
        'child_scholarship': ['子女獎助金'],
    }

    def __init__(self, main_or_reserved, dir_name, id_hidden=True, pattern_fill=None,
                 font=None, alignment=None):
        self._prefix = None
        self.pattern_fill = pattern_fill or PatternFill(
            start_color='F7F7F7', end_color='F7F7F7', fill_type='solid'
        )
        self.font = font or Font(bold=True)
        self.alignment = alignment or Alignment(horizontal='left', vertical='top', wrap_text=True)
        self.dir_name = dir_name
        self.main_or_reserved = main_or_reserved
        self.id_hidden = id_hidden
        self.__col_index = 1
        self.__row_index = 1
        self.__wb = openpyxl.Workbook()
        self.__sheet = self.__wb.active
        self.__crop_set = set()
        self.__set_column_width()
        self.__side = Side(border_style='double', color='000000')
        self.__border = Border(bottom=self.__side)
        self.num_format = lambda x: '#,##0' if type(x) == int and x > 0 else None

    @property
    def prefix(self):
        return self._prefix

    @prefix.setter
    def prefix(self, name):
        self._prefix = name

    @property
    def column_index(self):
        return self.__col_index

    @column_index.setter
    def column_index(self, i):
        self.__col_index += i

    @property
    def row_index(self):
        return self.__row_index

    @row_index.setter
    def row_index(self, i):
        self.__row_index += i

    def __set_column_width(self):
        width_dict = {
            'A': 13.29, 'B': 9.29, 'C': 12.29, 'D': 11.29, 'E': 10.29,
            'F': 10.29, 'G': 11.29, 'H': 10.29, 'I': 10.29, 'J': 7.79,
            'K': 7.79, 'L': 7.79, 'M': 7.79, 'N': 7.79, 'O': 7.79, 'P': 7.79,
        }

        width = list(map(lambda x: x * 1.054, width_dict.values()))

        for i in range(1, len(width) + 1):
            self.__sheet.column_dimensions[get_column_letter(i)].width = width[i - 1]
            # 應繳眷口數要隱藏, 調查員不需要, 但佩芬她們要看
            if get_column_letter(i) == 'N':
                self.__sheet.column_dimensions[get_column_letter(i)].hidden = True

        if not self.id_hidden:
            self.__sheet.column_dimensions['R'].width = 13.29

    def __set_column_merge(self, merge_list: list):
        """
        merge_list 每個元素是 tuple
        tuple[0]: 哪個欄位要合併, tuple[1]: 合併至哪個欄位
        e.g. ('A', 'B') 從'A'欄位併至'B', ('J', 'Q') 從'J'欄位併到'Q'

        :param merge_list:
        :return:
        """

        # 合併欄位
        for _tuple in merge_list:
            if type(_tuple) != tuple:
                raise TypeError(f"'{type(_tuple)}' is not a 'Tuple' object")

            start_col = column_index_from_string(_tuple[0])
            end_col = column_index_from_string(_tuple[1])
            self.__sheet.merge_cells(
                start_row=self.row_index, start_column=start_col,
                end_row=self.row_index, end_column=end_col
            )

    def __set_seprate_symbol(self):
        """
        每一戶資料之間的分隔線，使用 Excel 的 border
        :return:
        """
        # self.row_index = -2
        for col_index in range(1, column_index_from_string('Q') + 1):
            self.set_cell(row=self.row_index, col=col_index, border=self.__border)
        self.row_index = 1
        self.__sheet.cell(column=1, row=self.row_index).value = ''

    def set_title(self, field, pattern_fill=None, font=None, alignment=None,
                  new_line=True, passed_title=None, merge_list=None):
        """
        設定資料的 title, 會上底色以及字體加粗
        :param field:
        :param pattern_fill:
        :param font:
        :param alignment:
        :param new_line:
        :param passed_title:
        :param merge_list:
        :return:
        """
        next_col_index = None
        title = passed_title or self.titles[field]
        _pattern_fill = pattern_fill or self.pattern_fill
        _font = font or self.font
        # _alignment = alignment or self.alignment

        # 如果有要 merge 欄位
        if merge_list is not None:
            self.__set_column_merge(merge_list)

            # merge 過後的欄位, 只有合併欄位的第一欄能給值
            # e.g. A 與 B 欄合併, 只有 A 欄能給值, 若嘗試給 B 欄值會 Error
            # AB欄合併後，下一個能給值的欄位就是C欄, 若C欄合併到E欄
            # CDE欄只有C欄能給值，接著下一個能給值的欄位是F欄
            # merge_list 每個元素是 Tuple, tuple[0] 為合併欄位的起始欄位
            for val, col_index in zip(title, [column_index_from_string(t[0]) for t in merge_list]):
                if alignment is not None:
                    self.set_cell(self.row_index, col_index, val, _pattern_fill, _font, alignment)
                else:
                    self.set_cell(self.row_index, col_index, val, _pattern_fill, _font)
                next_col_index = col_index
        else:
            for col_index, val in enumerate(title, start=1):
                self.set_cell(self.row_index, col_index, val, _pattern_fill, _font, alignment)
                next_col_index = col_index

        if new_line:
            self.row_index = 1

        return next_col_index + 1

    def set_cell(self, row, col, value=None, pattern_fill=None, font=None, alignment=None,
                 new_line=False, number_format=None, border=None, wrap_text=True):
        """
        Excel 儲存格給值，若有要設定樣式則自行傳入
        :param row: 第N列
        :param col: 第M欄
        :param value: 儲存格的值
        :param pattern_fill:
        :param font: 字形
        :param alignment: 對齊
        :param new_line: 是否要換行
        :param number_format: 數字的格式 e.g. #,###,#### -> 1,000, 10,000, 100,000
        :param border: 邊線
        :param wrap_text:
        :return:
        """
        sheet = self.__sheet
        sheet.cell(row=row, column=col).value = value
        sheet.cell(row=row, column=col).fill = pattern_fill or PatternFill()
        sheet.cell(row=row, column=col).font = font or Font()
        sheet.cell(row=row, column=col).border = border or Border()
        sheet.cell(row=row, column=col).alignment = alignment or self.alignment
        sheet.cell(row=row, column=col).number_format = number_format if number_format else ''
        if new_line:
            self.row_index = 1

    def set_sample_info(self, field, data):
        """
        設定調查名冊戶長的基本資料
        :param field:
        :param data:
        :return:
        """
        merge_list = [
            ('A', 'B'), ('C', 'C'), ('D', 'D'), ('E', 'E'), ('F', 'G'), ('H', 'I'), ('J', 'Q')
        ]
        col_idex_list = [column_index_from_string(t[0]) for t in merge_list]
        self.set_title(field, merge_list=merge_list)

        # 因為 Excel 不需要寫入 ID 因此先 pop 出來
        data.pop('id')
        for val, col_index in zip(data.values(), col_idex_list):
            self.set_cell(row=self.row_index, col=col_index, value=val)

        self.__set_column_merge(merge_list)
        self.row_index = 2

    def set_household(self, field, data):
        """
        設定戶籍檔資料
        :param field:
        :param data:
        :return:
        """
        alignment = Alignment('left', 'top', wrap_text=True)
        # 如果不隱藏 id 則 title 要多加一個
        if not self.id_hidden:
            title = self.titles[field].copy()
            title.append('身分證字號')
            self.set_title(field, passed_title=title)
        else:
            self.set_title(field)

        for i, person in enumerate(data):
            # 同上，如果要顯示 id 就不 pop 出來
            person.pop() if self.id_hidden else ...
            if i > 0:
                self.row_index = 1
            for col_index, val in enumerate(person, start=1):
                self.set_cell(row=self.row_index, col=col_index, value=val, alignment=alignment,
                              number_format=self.num_format(val))

        self.row_index = 2

    def set_fallow_declare(self, field, data):
        """
        設定核定檔作物名稱
        :param field:
        :param data:
        :return:
        """
        if not data:
            return

        col_index = self.set_title(field, new_line=False)
        alignment = Alignment(horizontal='left', wrap_text=False)
        self.set_cell(row=self.row_index, col=col_index, value=data, alignment=alignment)
        self.row_index = 2

    def set_fallow_transfer_subsidy(self, field, data):
        """
        設定休耕轉作補貼
        :param field:
        :param data:
        :return:
        """
        if not data:
            return

        self.set_title(field)
        for index, _dict in enumerate(data, start=1):
            # 如果項目超過一個要換行
            if index > 1:
                self.row_index = 1

            _list = list(_dict.values())
            # 每一項前面要加上索引
            _list.insert(0, index)
            for col_index, val in enumerate(_list, start=2):
                self.set_cell(row=self.row_index, col=col_index, value=val,
                              number_format=self.num_format(val))

        self.row_index = 2

    def set_disaster_subsidy(self, field, data):
        """
        設定天然災害現金救助，同休耕轉作補貼設定
        :param field:
        :param data:
        :return:
        """
        if not data:
            return

        self.set_title(field)
        for index, _dict in enumerate(data, start=1):
            if index > 1:
                self.row_index = 1

            _list = list(_dict.values())
            _list.insert(0, index)
            for col_index, val in enumerate(_list, start=2):
                self.set_cell(row=self.row_index, col=col_index, value=val,
                              number_format=self.num_format(val))

        self.row_index = 2

    def set_livestock(self, field, data):
        """
        設定畜牧資料
        :param field:
        :param data:
        :return:
        """
        if not data:
            return

        # 因為欄位名稱會等於副產品名稱，所以當有副產品時才回去設定欄位
        # 因此要記住 row 與 col 的位置
        marked_col_index = self.set_title(field)
        marked_row_index = self.row_index - 1

        for field_name, data_list in data.items():
            set_field_name = False
            for index, _list in enumerate(data_list):
                if index > 0:
                    self.row_index = 1

                # 最後一個元素為副產品的數量，如果為 0 代表沒有副產品，於是就 pop 出來
                # 倒數第二個元素為副產品名稱，因為副產品數量為零也就不會有負產品名稱
                # 因此就以最後一個元素為依據，一次 pop 最後兩個元素
                if _list[-1] == 0:
                    _list.pop(-1)
                    _list.pop(-1)
                else:
                    product_field = _list.pop(-2)
                    self.set_cell(row=marked_row_index, col=marked_col_index, value=product_field,
                                  pattern_fill=self.pattern_fill, font=self.font)

                new_list = []
                # 每個畜牧場的名稱都只要顯示一次，設定過一次之後就不會再重複
                # 除非是不同的畜牧場名稱
                if not set_field_name:
                    new_list.append(field_name)
                    new_list.extend(_list)
                    set_field_name = True
                else:
                    new_list.append('')
                    new_list.extend(_list)

                for col_index, val in enumerate(new_list, start=1):
                    self.set_cell(row=self.row_index, col=col_index, value=val,
                                  number_format=self.num_format(val))

            self.row_index = 2

    def set_small_large_data(self, field, data):
        """
        設定小大資料
        :param field:
        :param data:
        :return:
        """
        if not data:
            return

        self.set_title(field)
        for index, _dict in enumerate(data, start=1):
            if index > 1:
                self.row_index = 1

            _list = list(_dict.values())
            for col_index, val in enumerate(_list, start=2):
                num_format = '#,###,###' if type(val) == int else None
                self.set_cell(row=self.row_index, col=col_index, value=val,
                              number_format=self.num_format(val))

        self.row_index = 2

    def set_crop_name(self, field, data):
        """
        設定作物名稱
        :param field:
        :param data:
        :return:
        """
        if not data:
            return

        alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
        col_index = self.set_title(field, new_line=False, alignment=alignment)
        self.set_cell(row=self.row_index, col=col_index, value=data, alignment=alignment, wrap_text=False)
        self.row_index = 2

    def set_child_scholarship(self, field, data):
        """
        設定子女獎助學金
        :param field:
        :param data:
        :return:
        """
        if not data:
            return

        col_index = self.set_title(field, new_line=False)
        alignment = Alignment(horizontal='left', wrap_text=False)
        self.set_cell(row=self.row_index, col=col_index, value=data, alignment=alignment)
        self.row_index = 2

    def generate_excel(self, data_dict):
        for field, data in data_dict.items():
            func = getattr(self, f"set_{field}", None)
            if func is None:
                raise NotImplementedError(f"{field} not implemented 'set_{field}'.")

            func(field, data)
        self.__set_seprate_symbol()

    def save(self):
        file_name = f"{self.prefix}_{self.main_or_reserved}公務資料" + '.xlsx'
        self.__wb.save(pjoin(self.dir_name, file_name))
        print(f"Fisished {file_name} ...", flush=True)

    @staticmethod
    def encrypt_excel():
        """
        加密 Excel
        :return:
        """
        password = src.encrypt
        for file in os.listdir(src.output_excel_dir):
            excel = win.gencache.EnsureDispatch('Excel.Application')
            wb = excel.Workbooks.Open(pjoin(src.output_excel_dir, file))
            wb.SaveAs(pjoin(os.getcwd(), src.output_encrypt_excel_dir, file), Password=password)
            excel.Application.Quit()


def get_result_data_frame(df_left_path, df_right_path, merge_key):
    left_df = pandas.read_excel(df_left_path, dtype={'農戶編號': 'str', '電話': 'str', })
    right_df = pandas.read_excel(df_right_path, dtype={'農戶編號': 'str', '電話': 'str', })
    result = pandas.merge(left_df, right_df, how='left', on=merge_key, validate='1:1')
    return result


def task(_list):
    data_frame, main_or_reserved, prefix, dir_name, is_hidden = _list
    # file_name = f"{main_or_reserved}_公務資料.json"
    file_name = 'test.json'
    json_data = read_json_file(pjoin(src.output_json_dir, file_name))
    handler = ExcelHandler(main_or_reserved, dir_name, is_hidden)
    handler.prefix = data_frame.iloc[0][prefix]

    for i in range(len(data_frame.index)):
        series = data_frame.iloc[i]
        _dict = json_data.get(series['農戶編號'])

        if _dict is None:
            print(f"\n{series['農戶編號']} data is None.", flush=True)
            continue
        handler.generate_excel(_dict)
    handler.save()


if __name__ == '__main__':
    merge_key = '農戶編號'
    group_by_key = '調查員'
    _input = input('0 -> 主選\n1 -> 備選\nChoicing mode:')
    df_list = None
    suffix = datetime.now().strftime('%Y%m%d_%H%M%S')
    flag = not bool(int(_input))
    id_hidden = bool(int(input('0 -> 是\n1 -> 否\n是否顯示 ID:')))

    if flag:
        _input = input('0 -> 依調查員\n1 -> 依縣市\nChoicing mode:')
        if _input == '0':
            dir_name = pjoin(src.output_excel_dir, f"主選公務資料(依調查員)_{suffix}")
            mkdir(dir_name)
            result = get_result_data_frame(
                pjoin(src.input_excel_dir, '主選名冊.xlsx'),
                pjoin(src.input_excel_dir, '調查名冊.xlsx'),
                merge_key
            )
            groups = result.groupby(group_by_key)
            df_list = [
                [groups.get_group(inv_name).sort_values('地址'), '主選', '調查員', dir_name, id_hidden]
                for inv_name in groups.groups
            ]
        elif _input == '1':
            dir_name = pjoin(src.output_excel_dir, f"主選公務資料(依縣市3套)_{suffix}")
            mkdir(dir_name)
            df = pandas.read_excel(pjoin(src.input_excel_dir, '調查名冊.xlsx'), dtype={
                '電話': 'str',
                '農戶編號': 'str',
            })

            # 濾出備選(沒有 '*' 為備選) 並且連結編號最後一碼 <= 3(因為總共有 1~6)
            # 且農戶編號沒有在家庭收支調查裡面
            df = df[(df['主備標記'] == '*')]

            df['戶長姓名'] = df['戶長姓名'].str.replace('\u3000', '').replace(' ', '')
            df['縣市'] = df['地址'].str.replace('.*?(..[縣市]).*', r'\1', regex=True)
            groups = df.groupby('縣市')
            df_list = [
                [groups.get_group(county).sort_values('地址'), '主選', '縣市', dir_name, id_hidden]
                for county in groups.groups
            ]
    else:
        _input = input('0 -> 依調查員3套\n1 -> 依縣市3套\n2 -> 依縣市6套\nChoicing mode:')
        if _input == '0':
            def sort_data_frame(df) -> pandas.DataFrame:
                parser = lambda x: str(x).rjust(5, '0')
                df['連結編號'] = df['連結編號'].apply(parser)
                # 連結編號前四碼為群組依據
                # e.g. 00010 00011 00012 -> 前四碼都是 0001
                # 10121 10122 10123 -> 1012
                df['groupBy'] = df['連結編號'].str[:4]
                # 先只取第一套出來排序，之後再插入第2、3套
                df_first = df[df['連結編號'].str.match(r'.*1$')].sort_values('地址')
                groups = df.groupby('groupBy')
                sorted_df = pandas.DataFrame()
                for i in range(len(df_first.index)):
                    sorted_df = sorted_df.append(groups.get_group(df_first.iloc[i]['groupBy']))
                return sorted_df

            dir_name = pjoin(src.output_excel_dir, f"備選公務資料(依調查員)_{suffix}")
            mkdir(dir_name)
            result = get_result_data_frame(
                pjoin(src.input_excel_dir, '備選3套名冊.xlsx'),
                pjoin(src.input_excel_dir, '調查名冊.xlsx'),
                merge_key
            )
            groups = result.groupby(group_by_key)
            df_list = [
                [sort_data_frame(groups.get_group(inv_name)), '備選', '調查員', dir_name, id_hidden]
                for inv_name in groups.groups
            ]
        elif _input == '1' or _input == '2':
            letter = '3' if _input == '1' else '6'
            dir_name = pjoin(src.output_excel_dir, f"備選公務資料(依縣市{letter}套)_{suffix}")
            mkdir(dir_name)
            df = pandas.read_excel(pjoin(src.input_excel_dir, '調查名冊.xlsx'), dtype={
                '電話': 'str',
                '農戶編號': 'str',
            })

            # 濾出備選(沒有 '*' 為備選) 並且連結編號最後一碼 <= 3(因為總共有 1~6)
            df = df[(df['主備標記'] != '*') & (df['連結編號'] % 10 <= 3)]\
                if _input == '1' else df[(df['主備標記'] != '*')]

            df['戶長姓名'] = df['戶長姓名'].str.replace('\u3000', '').replace(' ', '')
            df['縣市'] = df['地址'].str.replace('.*?(..[縣市]).*', r'\1', regex=True)
            groups = df.groupby('縣市')
            df_list = [
                [groups.get_group(county).sort_values('連結編號'), '備選', '縣市', dir_name, id_hidden]
                for county in groups.groups
            ]

    @timer
    def run_tasks():
        with multiprocessing.Pool() as p:
            var = [p.map(task, df_list)]


    run_tasks()
