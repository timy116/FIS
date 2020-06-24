import functools
import multiprocessing
import time
import openpyxl
import pandas

from collections import OrderedDict
from datetime import datetime
from os.path import join as pjoin
from os import mkdir
from openpyxl.styles import Alignment, Side, Border
from openpyxl.utils import get_column_letter, column_index_from_string
from pprint import pprint
from resources import src
from main import ConcreteGenerateOfficialData


EXCEPT_LIST = ConcreteGenerateOfficialData.EXCEPT_LIST


def timer(func):
    @functools.wraps(func)
    def wrapper_timer(*args, **kwargs):
        start_time = time.perf_counter()
        func(*args, **kwargs)
        end_time = time.perf_counter()
        run_time = end_time - start_time
        m, s = divmod(run_time, 60)

        msg = f"execute time: {int(m)} min, {s:.1f} sec"
        print(msg)

    return wrapper_timer


class ExcelHandler:
    roc = datetime.today().year - 1911
    titles = {
        'three_lines': [
            '{roc}年主力農家所得調查樣本名冊─{main_or_reserved}',
            '本頁已完成調查戶數：_____',
            '失敗戶請填寫失敗訪視紀錄表',
            '',
        ],
        'fifth_line': ['序號', '樣本套號 ', '農戶編號', '連結編號 ', '戶長姓名', '電話 ',
                       '地址 ', '層別 ', '經營種類 ', '可耕作地面積', '成功打勾', '有無外僱'],
    }

    def __init__(self, main_or_reserved, dir_name, alignment=None):
        self.alignment = alignment or Alignment(horizontal='left', wrap_text=True)
        self.main_or_reserved = main_or_reserved
        self.dir_name = dir_name
        self._prefix = None
        self.__col_index = 1
        self.__row_index = 1
        self.__wb = openpyxl.Workbook()
        self.__sheet = self.__wb.active
        self.__crop_set = set()
        self.__set_column_width()
        self.__side = Side(style='medium')
        self.__border = Border(top=self.__side, right=self.__side, bottom=self.__side, left=self.__side)
        self.__set_three_lines_title()

    @property
    def prefix(self):
        return self._prefix

    @prefix.setter
    def prefix(self, name):
        self._prefix = name

    @property
    def column_index(self):
        return self.__col_index

    @column_index.setter
    def column_index(self, i):
        if i == -1:
            self.__col_index = 1
        else:
            self.__col_index += i

    @property
    def row_index(self):
        return self.__row_index

    @row_index.setter
    def row_index(self, i):
        self.__row_index += i

    def __set_column_width(self):
        width = list(
            map(lambda x: x * 1.13, [5.29, 5.29, 13.29, 9.29, 11.29, 12.29, 50.29, 4.29, 10.29, 20.29, 5.29, 5.29])
        )
        for i in range(1, len(width) + 1):
            self.__sheet.column_dimensions[get_column_letter(i)].width = width[i - 1]
        self.row_index = 1

    def __set_three_lines_title(self):
        """
        設定前三行格式(第四行為空行, 但會給 border)
        :return:
        """
        titles = ExcelHandler.titles['three_lines']
        alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for index, title in enumerate(titles, start=1):
            sheet = self.__sheet
            # 合併欄位, 從 A 欄併到 L 欄
            sheet.merge_cells(
                start_row=index,
                start_column=self.column_index,
                end_row=index,
                end_column=column_index_from_string('L')
            )

            if index == 1:
                title = title.format(roc=self.roc, main_or_reserved=self.main_or_reserved)
            elif index == 3:
                alignment = Alignment(horizontal='right')
            elif index == 4:
                for i in range(1, 13):
                    self.set_cell(index, i, border=self.__border)

            self.set_cell(index, self.column_index, title, alignment=alignment, new_line=True)

    def set_title(self, field, alignment=None, border=None, new_line=True, passed_title=None):
        next_col_index = None
        title = passed_title or self.titles[field]
        _alignment = alignment or self.alignment
        _border = border or None

        for col_index, val in enumerate(title, start=1):
            self.set_cell(self.row_index, col_index, val, _alignment, border=border)
            next_col_index = col_index

        if new_line:
            self.row_index = 1

        return next_col_index + 1

    def set_cell(self, row, col, value=None, alignment=None,
                 new_line=False, number_format=None, border=None):
        sheet = self.__sheet

        if value is not None:
            sheet.cell(row=row, column=col).value = value

        if alignment is not None:
            sheet.cell(row=row, column=col).alignment = alignment
        else:
            sheet.cell(row=row, column=col).alignment = self.alignment

        if number_format:
            sheet.cell(row=row, column=col).number_format = number_format
        if border is not None:
            sheet.cell(row=row, column=col).border = border
        if new_line:
            self.row_index = 1

    def generate_excel(self, data_frame):
        update_df = ExcelHandler.read_update_file()
        data_frame = pandas\
            .merge(data_frame, update_df, how='left', on='ID')\
            .fillna('')

        alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        # 退一行
        self.row_index = -1
        self.set_title('fifth_line', alignment=alignment, border=self.__border)

        for i in range(len(data_frame.index)):
            # 設定 row height, 否則看起來會很擠
            self.__sheet.row_dimensions[self.row_index].height = 1.95 * 17.153
            _dict = OrderedDict()
            series = data_frame.iloc[i]

            _dict['序號'] = i + 1
            _dict['樣本套號'] = series['樣本套號']
            _dict['農戶編號'] = series['農戶編號']
            _dict['連結編號'] = str(series['連結編號']).rjust(5, '0')
            _dict['戶長姓名'] = series['戶長姓名']
            _dict['電話'] = '0' + series['電話'] if series['電話'] and not series['電話'].startswith('0')\
                else series['電話']
            _dict['地址'] = series['地址']
            _dict['層別'] = series['層別']
            _dict['經營種類'] = series['主要經營型態']
            _dict['可耕作地面積'] = series['可耕作地面積']
            _dict['成功打勾'] = ''
            _dict['有無外僱'] = series['有無外僱']

            # 對應地址, 若一樣則更新電話或手機或者對應戶長姓名
            if series['地址'] == series['addr'] or series['戶長姓名'] == series['name']:
                # 姓名一樣則更新地址
                if series['戶長姓名'] == series['name']:
                    _dict['地址'] = series['addr']
                if series['tel']:
                    _dict['電話'] = series['tel']
                if series['phone']:
                    _dict['電話'] += '/\n' + series['phone']

            for col, val in enumerate(_dict.values(), start=1):
                self.set_cell(self.row_index, col, val, border=self.__border)
            self.row_index = 1

    def save(self):
        file_name = f"{self.prefix}_{self.main_or_reserved}樣本名冊" + '.xlsx'
        self.__wb.save(
            pjoin(self.dir_name, file_name)
        )
        print(f"Fisished {file_name} ...", flush=True)

    @staticmethod
    def read_update_file():
        _dir = src.input_excel_dir
        file = 'update.xlsx'
        df = pandas.read_excel(pjoin(_dir, file), dtype={
            '連絡電話': 'str',
            '手機號碼': 'str',
        })\
            .rename(columns={
                '戶長姓名': 'name',
                '連絡電話': 'tel',
                '手機號碼': 'phone',
                '地址': 'addr',
            })

        return df


def get_result_data_frame(df_left_path, df_right_path, merge_key):
    left_df = pandas.read_excel(df_left_path, dtype={'農戶編號': 'str', '電話': 'str', })
    right_df = pandas.read_excel(df_right_path, dtype={'農戶編號': 'str', '電話': 'str', })
    result = pandas.merge(left_df, right_df, how='left', on=merge_key, validate='1:1')
    return result


def task(_list):
    data_frame, main_or_reserved, prefix, dir_name = _list
    handler = ExcelHandler(main_or_reserved, dir_name)
    handler.prefix = data_frame.iloc[0][prefix]
    handler.generate_excel(data_frame)
    handler.save()


if __name__ == '__main__':
    merge_key = '農戶編號'
    group_by_key = '調查員'
    _input = input('0 -> 主選\n1 -> 備選3套(依調查員)\n2 -> 備選3套(依縣市):\nChoicing mode:')
    df_list = None
    suffix = datetime.now().strftime('%Y%m%d_%H%M%S')

    if _input == '0':
        dir_name = pjoin(src.output_excel_dir, f"主選樣本名冊(依調查員)_{suffix}")
        mkdir(dir_name)
        result = get_result_data_frame(
            pjoin(src.input_excel_dir, '主選名冊.xlsx'),
            pjoin(src.input_excel_dir, '調查名冊.xlsx'),
            merge_key
        )
        groups = result.groupby(group_by_key)
        df_list = [
            [groups.get_group(county).sort_values('地址'), '主選', '調查員', dir_name]
            for county in groups.groups
        ]

    elif _input == '1':
        dir_name = pjoin(src.output_excel_dir, f"備選樣本名冊(依調查員3套)_{suffix}")
        mkdir(dir_name)
        result = get_result_data_frame(
            pjoin(src.input_excel_dir, '備選3套名冊.xlsx'),
            pjoin(src.input_excel_dir, '調查名冊.xlsx'),
            merge_key
        )
        result = result[~result['農戶編號'].isin(EXCEPT_LIST)]
        groups = result.groupby(group_by_key)
        df_list = [
            [groups.get_group(county).sort_values('連結編號'), '備選', '調查員', dir_name]
            for county in groups.groups
        ]

    else:
        dir_name = pjoin(src.output_excel_dir, f"備選樣本名冊(依縣市3套)_{suffix}")
        mkdir(dir_name)
        df = pandas.read_excel(pjoin(src.input_excel_dir, '調查名冊.xlsx'), dtype={
            '電話': 'str',
            '農戶編號': 'str',
        })

        df = df[
            (df['主備標記'] != '*')
            & (df['連結編號'] % 10 <= 3)
            & (~df['農戶編號'].isin(EXCEPT_LIST))
        ]

        df['戶長姓名'] = df['戶長姓名'].str.replace('\u3000', '').replace(' ', '')
        df['縣市'] = df['地址'].str.replace('.*?(..[縣市]).*', r'\1', regex=True)
        groups = df.groupby('縣市')
        df_list = [
            [groups.get_group(county).sort_values('連結編號'), '備選', '縣市', dir_name]
            for county in groups.groups
        ]

    @timer
    def run_tasks():
        with multiprocessing.Pool() as p:
            var = [p.map(task, df_list)]

    run_tasks()
